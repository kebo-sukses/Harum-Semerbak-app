# -*- coding: utf-8 -*-
"""
excel_template.py — Generator template Excel untuk import massal
                     & konversi Bahasa Indonesia → aksara Mandarin.

Modul ini menyediakan:
  1. Daftar dropdown Panggilan & Dari dalam Bahasa Indonesia.
  2. Fungsi konversi Indonesia → Mandarin (otomatis saat import).
  3. Generator file Excel (.xlsx) template siap pakai.
"""

from __future__ import annotations


# ============================================================
# Angka Mandarin (duplikasi dari main.py agar modul mandiri)
# ============================================================
_MANDARIN_NUMBERS: dict[int, str] = {
    1: "長", 2: "二", 3: "三", 4: "四", 5: "五",
    6: "六", 7: "七", 8: "八", 9: "九", 10: "十",
}
_MANDARIN_NUMBERS_ADIK: dict[int, str] = {
    1: "大", 2: "二", 3: "三", 4: "四", 5: "五",
    6: "六", 7: "七", 8: "八", 9: "九", 10: "十",
}
_DARI_ANAK_NUMBERS: dict[int, str] = {
    1: "長", 2: "次", 3: "三", 4: "四", 5: "五",
    6: "六", 7: "七", 8: "八", 9: "九", 10: "十",
}

_SUFFIX = " 敬奉 叩首"


# ============================================================
# PANGGILAN: Indonesian → (Mandarin, Keluarga)
# ============================================================
_PANGGILAN_TO_MANDARIN: dict[str, str] = {}
_PANGGILAN_TO_KELUARGA: dict[str, str] = {}
PANGGILAN_INDO_LIST: list[str] = []

# --- Item non-bernomor ---
_PANGGILAN_BASE: list[tuple[str, str, str]] = [
    ("Ayah", "父親", "Ayah Kandung"),
    ("Kakek (Pihak Ayah)", "祖父", "Kakek Kandung"),
    ("Nenek (Pihak Ayah)", "祖母", "Nenek Kandung"),
    ("Kakak Laki-laki Ayah", "伯父", "Paman (Kakak Ayah)"),
    ("Istri Kakak Laki-laki Ayah", "伯母", "Istri Paman (Kakak Ayah)"),
    ("Adik Laki-laki Ayah", "叔父", "Paman (Adik Ayah)"),
    ("Istri Adik Laki-laki Ayah", "嬸母", "Istri Paman (Adik Ayah)"),
    ("Saudara Perempuan Ayah (Bibi)", "姑母", "Bibi (Saudara Ayah)"),
    ("Suami Saudara Perempuan Ayah", "姑丈", "Suami Bibi (Saudara Ayah)"),
    ("Kakak Dari Kakek", "太伯祖考", "Kakak Kakek"),
    ("Ibu", "母親", "Ibu Kandung"),
    ("Kakek (Pihak Ibu)", "外祖父", "Kakek Luar"),
    ("Nenek (Pihak Ibu)", "外祖母", "Nenek Luar"),
    ("Saudara Laki-laki Ibu", "舅父", "Paman (Saudara Ibu)"),
    ("Istri Saudara Laki-laki Ibu", "舅母", "Istri Paman (Saudara Ibu)"),
    ("Saudara Perempuan Ibu", "姨母", "Bibi (Saudara Ibu)"),
    ("Suami Saudara Perempuan Ibu", "姨丈", "Suami Bibi (Saudara Ibu)"),
    ("Kakak/Adik Ipar Laki-laki", "姊夫", "Ipar Laki-laki"),
    ("Saudara Laki-laki Istri", "内兄", "Saudara Laki-laki Istri"),
    ("Suami Saudara Perempuan Istri", "襟兄", "Suami Saudara Perempuan Istri"),
    ("Kakak/Adik Ipar Perempuan", "嫂嫂", "Ipar Perempuan"),
    ("Saudara Perempuan Istri", "内姊", "Saudara Perempuan Istri"),
    ("Istri Saudara Laki-laki Istri", "内嫂", "Istri Saudara Laki-laki Istri"),
    ("Ayah Angkat", "養父", "Ayah Angkat"),
    ("Ibu Angkat", "養母", "Ibu Angkat"),
    ("Kakek Angkat (Ayah)", "養祖父", "Kakek Angkat"),
    ("Nenek Angkat (Ayah)", "養祖母", "Nenek Angkat"),
    ("Kakek Angkat (Ibu)", "養外祖父", "Kakek Angkat Luar"),
    ("Nenek Angkat (Ibu)", "養外祖母", "Nenek Angkat Luar"),
    ("Menantu Laki-laki", "女婿", "Menantu Laki-laki"),
    ("Menantu Perempuan", "媳妇", "Menantu Perempuan"),
]

for _indo, _mandarin, _keluarga in _PANGGILAN_BASE:
    PANGGILAN_INDO_LIST.append(_indo)
    _PANGGILAN_TO_MANDARIN[_indo] = _mandarin
    _PANGGILAN_TO_KELUARGA[_indo] = _keluarga

# --- Item bernomor (4 kategori × 10 nomor = 40 item) ---
_PANGGILAN_NUMBERED_TEMPLATES: list[tuple[str, str, str, bool]] = [
    # (base_indo, template_mandarin, keluarga_base, is_adik)
    ("Kakak Laki-laki", "亡胞__兄", "Kakak Laki-laki", False),
    ("Kakak Perempuan", "亡胞__姊", "Kakak Perempuan", False),
    ("Adik Laki-laki", "亡胞__弟", "Adik Laki-laki", True),
    ("Adik Perempuan", "亡胞__妹", "Adik Perempuan", True),
]

for _base, _template, _kel_base, _is_adik in _PANGGILAN_NUMBERED_TEMPLATES:
    _lookup = _MANDARIN_NUMBERS_ADIK if _is_adik else _MANDARIN_NUMBERS
    for _num in range(1, 11):
        _key = f"{_base} Ke-{_num}"
        _cn = _lookup[_num]
        _mand = _template.replace("__", _cn)
        PANGGILAN_INDO_LIST.append(_key)
        _PANGGILAN_TO_MANDARIN[_key] = _mand
        _PANGGILAN_TO_KELUARGA[_key] = f"{_kel_base} {_num}"


# ============================================================
# DARI: Indonesian → Mandarin (sudah termasuk suffix)
# ============================================================
_DARI_TO_MANDARIN: dict[str, str] = {}
DARI_INDO_LIST: list[str] = []

# --- Non-gendered, non-numbered ---
_DARI_SIMPLE: list[tuple[str, str]] = [
    ("Anak Lk dan Pr", "众孝眷"),
]

for _indo, _mand in _DARI_SIMPLE:
    DARI_INDO_LIST.append(_indo)
    _DARI_TO_MANDARIN[_indo] = f"{_mand}{_SUFFIX}"

# --- Gendered, non-numbered ---
_DARI_GENDERED_BASE: list[tuple[str, str, str]] = [
    # (base_indo, mandarin_L, mandarin_P)
    ("Anak Bungsu", "孝幼子", "孝幼女"),
    ("Anak Angkat", "孝養子", "孝養女"),
    ("Anak Tiri", "孝繼子", "孝繼女"),
    ("Menantu", "孝女婿", "孝兒媳"),
    ("Cucu Kandung", "孝孫", "孝孫女"),
    ("Cucu Luar", "孝外孫", "孝外孫女"),
    ("Keponakan (Sdr Laki-laki)", "孝姪", "孝姪女"),
    ("Keponakan (Sdr Perempuan)", "孝外甥", "孝外甥女"),
    ("Saudara Ipar (Adik)", "孝內弟", "孝內姊"),
    ("Saudara Ipar (Kakak)", "内兄", "内妹"),
]

for _base, _laki, _perempuan in _DARI_GENDERED_BASE:
    for _gender, _m in [("Laki-laki", _laki), ("Perempuan", _perempuan)]:
        _key = f"{_base} ({_gender})"
        DARI_INDO_LIST.append(_key)
        _DARI_TO_MANDARIN[_key] = f"{_m}{_SUFFIX}"

# --- Gendered + bernomor (3 kategori × 10 × 2 = 60 item) ---
_DARI_NUMBERED_TEMPLATES: list[tuple[str, str, str, str]] = [
    # (base_indo, template_L, template_P, number_type)
    ("Anak", "孝__子", "孝__女", "anak"),
    ("Adik", "愚__弟", "愚__妹", "adik"),
    ("Kakak", "胞__兄", "胞__姊", "kakak"),
]

for _base, _tmpl_l, _tmpl_p, _ntype in _DARI_NUMBERED_TEMPLATES:
    if _ntype == "anak":
        _num_lookup = _DARI_ANAK_NUMBERS
    elif _ntype == "adik":
        _num_lookup = _MANDARIN_NUMBERS_ADIK
    else:
        _num_lookup = _MANDARIN_NUMBERS

    for _num in range(1, 11):
        _cn = _num_lookup[_num]
        for _gender, _tmpl in [("Laki-laki", _tmpl_l), ("Perempuan", _tmpl_p)]:
            _key = f"{_base} Ke-{_num} ({_gender})"
            _mand = _tmpl.replace("__", _cn)
            DARI_INDO_LIST.append(_key)
            _DARI_TO_MANDARIN[_key] = f"{_mand}{_SUFFIX}"

# --- Frasa umum (tanpa suffix, tanpa gender) ---
_DARI_FRASA_UMUM: list[tuple[str, str]] = [
    ("众孝眷 偕 合家敬奉", "众孝眷 偕 合家敬奉"),
    ("孝子贤孙 偕 合家敬奉", "孝子贤孙 偕 合家敬奉"),
    ("合家敬奉 叩首", "合家敬奉 叩首"),
]

for _indo, _mand in _DARI_FRASA_UMUM:
    DARI_INDO_LIST.append(_indo)
    _DARI_TO_MANDARIN[_indo] = _mand  # tanpa suffix tambahan


# ============================================================
# Fungsi Konversi (dipakai oleh import_from_excel)
# ============================================================
def convert_panggilan(indo_text: str) -> tuple[str, str] | None:
    """Konversi teks Panggilan Indonesia → (mandarin, keluarga).

    Args:
        indo_text: Teks Indonesia dari dropdown Excel.

    Returns:
        (mandarin, keluarga) jika ditemukan, None jika tidak.
    """
    mandarin = _PANGGILAN_TO_MANDARIN.get(indo_text)
    if mandarin is None:
        return None
    keluarga = _PANGGILAN_TO_KELUARGA.get(indo_text, "")
    return (mandarin, keluarga)


def convert_dari(indo_text: str) -> str | None:
    """Konversi teks Dari Indonesia → aksara Mandarin (termasuk suffix).

    Args:
        indo_text: Teks Indonesia dari dropdown Excel.

    Returns:
        Aksara Mandarin lengkap jika ditemukan, None jika tidak.
    """
    return _DARI_TO_MANDARIN.get(indo_text)


# ============================================================
# Generator Template Excel
# ============================================================
def generate_template(output_path: str, num_rows: int = 20) -> str:
    """Buat file template Excel untuk import massal.

    Template berisi:
      - Sheet "Data Import" : tabel kosong dengan dropdown Panggilan & Dari.
      - Sheet "Referensi"   : daftar opsi dropdown (jangan diedit).

    Args:
        output_path: Path tujuan file .xlsx.
        num_rows:    Jumlah baris data (default 20).

    Returns:
        Path file yang berhasil dibuat.

    Raises:
        RuntimeError: Jika openpyxl belum terinstal atau gagal menulis.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, Alignment, PatternFill, Border, Side,
        )
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Library openpyxl belum terinstal.\n"
            "Jalankan: pip install openpyxl"
        ) from exc

    wb = Workbook()

    # ── Sheet 1: Data Import ─────────────────────────────────
    ws = wb.active
    ws.title = "Data Import"
    ws.sheet_properties.tabColor = "4472C4"

    # Style
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell_font = Font(name="Segoe UI", size=11)
    mandarin_font = Font(name="SimSun", size=12)

    # Header row
    headers = [
        ("NO", 5),
        ("NAMA PEMESAN", 28),
        ("PANGGILAN", 38),
        ("ATAS NAMA MANDARIN", 26),
        ("PENYEBUTAN", 22),
        ("DARI", 38),
        ("KETERANGAN", 22),
    ]
    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows (kosong, hanya nomor di kolom A)
    data_align = Alignment(vertical="center")
    for row_idx in range(2, num_rows + 2):
        # NO
        cell_no = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        cell_no.font = cell_font
        cell_no.alignment = Alignment(horizontal="center", vertical="center")
        cell_no.border = thin_border

        # Kolom B-G: border + font default
        for col_idx in range(2, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = data_align
            if col_idx == 4:  # Atas Nama Mandarin → font SimSun
                cell.font = mandarin_font
            else:
                cell.font = cell_font

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Sheet 2: Referensi (daftar dropdown) ─────────────────
    ws_ref = wb.create_sheet("Referensi")
    ws_ref.sheet_properties.tabColor = "FFC000"

    # Header
    ref_header_fill = PatternFill("solid", fgColor="FFC000")
    for col_idx, title in enumerate(["PANGGILAN", "DARI"], start=1):
        cell = ws_ref.cell(row=1, column=col_idx, value=title)
        cell.font = Font(name="Segoe UI", size=10, bold=True)
        cell.fill = ref_header_fill

    # Isi daftar Panggilan di kolom A
    for i, name in enumerate(PANGGILAN_INDO_LIST, start=2):
        ws_ref.cell(row=i, column=1, value=name).font = cell_font

    # Isi daftar Dari di kolom B
    for i, name in enumerate(DARI_INDO_LIST, start=2):
        ws_ref.cell(row=i, column=2, value=name).font = cell_font

    ws_ref.column_dimensions["A"].width = 40
    ws_ref.column_dimensions["B"].width = 42

    # ── Data Validation: dropdown Panggilan (kolom C) ────────
    pang_count = len(PANGGILAN_INDO_LIST) + 1  # +1 karena row 1 = header
    dv_panggilan = DataValidation(
        type="list",
        formula1=f"Referensi!$A$2:$A${pang_count}",
        allow_blank=True,
    )
    dv_panggilan.error = "Pilih panggilan dari daftar di sheet Referensi."
    dv_panggilan.errorTitle = "Panggilan Tidak Valid"
    dv_panggilan.prompt = "Pilih panggilan dari daftar"
    dv_panggilan.promptTitle = "Panggilan"
    ws.add_data_validation(dv_panggilan)
    dv_panggilan.add(f"C2:C{num_rows + 1}")

    # ── Data Validation: dropdown Dari (kolom F) ─────────────
    dari_count = len(DARI_INDO_LIST) + 1
    dv_dari = DataValidation(
        type="list",
        formula1=f"Referensi!$B$2:$B${dari_count}",
        allow_blank=True,
    )
    dv_dari.error = "Pilih 'Dari' dari daftar di sheet Referensi."
    dv_dari.errorTitle = "Dari Tidak Valid"
    dv_dari.prompt = "Pilih hubungan 'Dari' dari daftar"
    dv_dari.promptTitle = "Dari"
    ws.add_data_validation(dv_dari)
    dv_dari.add(f"F2:F{num_rows + 1}")

    # ── Simpan file ──────────────────────────────────────────
    try:
        wb.save(output_path)
    except OSError as exc:
        raise RuntimeError(f"Gagal menyimpan template: {exc}") from exc

    return output_path
