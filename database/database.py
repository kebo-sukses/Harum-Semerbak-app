# -*- coding: utf-8 -*-
"""
database.py — Modul SQLite untuk menyimpan data formulir ritual.

Schema 2-tabel (relasional):

  TABEL `orders` — Data pemesan unik
    id          : INTEGER PRIMARY KEY
    uuid        : TEXT UNIQUE
    nama        : TEXT (nama Indonesia: Ajon Bengkel, Rudi Jajan, Lili)
    created_at  : TEXT

  TABEL `ritual_items` — Detail formulir ritual (many-to-one → orders)
    id             : INTEGER PRIMARY KEY
    uuid           : TEXT UNIQUE
    order_id       : INTEGER FK → orders.id
    panggilan      : TEXT (sebutan Mandarin: 母親許門, 父親, 祖父)
    nama_mandarin  : TEXT (nama Mandarin mendiang: 梁氏橋玉)
    penyebutan     : TEXT (romanisasi/Hokkien: Nio Kiaw Gek)
    dari           : TEXT (pengirim/hubungan: 孝男, 外孫女敬奉)
    keluarga       : TEXT (relasi Indonesia: Ibu Kandung, dll)
    keterangan     : TEXT (info tambahan: 合家敬奉, dll)
    tahun_lunar    : TEXT
    bulan_lunar    : TEXT
    hari_lunar     : TEXT
    created_at     : TEXT

Keuntungan:
  - Nama pemesan yang sama (misal "Ajon Bengkel") hanya 1 baris di orders.
  - Semua formulir milik pemesan tsb di-link via order_id.
  - UI Treeview bisa menampilkan data terkelompok per pemesan.
"""

import sqlite3
import uuid
import os
from datetime import datetime


# ============================================================
# Path default database — disimpan di folder /database
# ============================================================
_DB_DIR = os.path.dirname(os.path.abspath(__file__))
_DB_PATH = os.path.join(_DB_DIR, "ritual_forms.db")


def _get_connection(db_path: str = _DB_PATH) -> sqlite3.Connection:
    """
    Membuka koneksi ke database SQLite.
    Mengaktifkan WAL mode & foreign keys.

    Args:
        db_path: Path absolut ke file .db

    Returns:
        sqlite3.Connection
    """
    try:
        conn = sqlite3.connect(db_path)
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA foreign_keys=ON;")
        conn.row_factory = sqlite3.Row
        return conn
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal membuka database: {e}") from e


# ============================================================
# Inisialisasi & Migrasi
# ============================================================
def init_db(db_path: str = _DB_PATH) -> None:
    """
    Membuat tabel `orders` dan `ritual_items` jika belum ada.
    Jika tabel lama `ritual_forms` masih ada, otomatis migrasi datanya
    ke skema baru, lalu rename tabel lama.

    Args:
        db_path: Path absolut ke file .db
    """
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()

        # --- Buat tabel baru ---
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid        TEXT    NOT NULL UNIQUE,
                nama        TEXT    NOT NULL DEFAULT '',
                created_at  TEXT    NOT NULL
            );
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ritual_items (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid           TEXT    NOT NULL UNIQUE,
                order_id       INTEGER NOT NULL
                                   REFERENCES orders(id) ON DELETE CASCADE,
                panggilan      TEXT    NOT NULL DEFAULT '',
                nama_mandarin  TEXT    NOT NULL DEFAULT '',
                penyebutan     TEXT    DEFAULT '',
                dari           TEXT    NOT NULL DEFAULT '',
                keluarga       TEXT    DEFAULT '',
                keterangan     TEXT    DEFAULT '',
                tahun_lunar    TEXT    DEFAULT '',
                bulan_lunar    TEXT    DEFAULT '',
                hari_lunar     TEXT    DEFAULT '',
                created_at     TEXT    NOT NULL
            );
        """)

        # Index untuk pencarian cepat per order
        cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_items_order_id
            ON ritual_items(order_id);
        """)

        conn.commit()

        # --- Migrasi data lama jika tabel `ritual_forms` masih ada ---
        _migrate_legacy(conn)

    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal membuat tabel: {e}") from e
    finally:
        conn.close()


def _migrate_legacy(conn: sqlite3.Connection) -> None:
    """
    Migrasi data dari tabel lama `ritual_forms` ke skema baru
    (`orders` + `ritual_items`).  Setelah migrasi, tabel lama
    di-rename menjadi `ritual_forms_backup`.
    """
    cursor = conn.cursor()

    # Cek apakah tabel lama masih ada
    cursor.execute(
        "SELECT name FROM sqlite_master "
        "WHERE type='table' AND name='ritual_forms';"
    )
    if cursor.fetchone() is None:
        return  # Tidak ada tabel lama — skip

    # Ambil semua data lama
    cursor.execute("SELECT * FROM ritual_forms ORDER BY created_at;")
    old_rows = cursor.fetchall()
    if not old_rows:
        # Tabel lama kosong — langsung rename
        cursor.execute("ALTER TABLE ritual_forms RENAME TO ritual_forms_backup;")
        conn.commit()
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Kelompokkan per nama (case-insensitive, strip whitespace)
    orders_map: dict[str, int] = {}  # nama_lower → order_id

    for row in old_rows:
        row_dict = dict(row)
        nama_raw = (row_dict.get("nama") or "").strip()
        nama_key = nama_raw.lower() if nama_raw else ""

        # Buat order jika belum ada
        if nama_key not in orders_map:
            order_uuid = str(uuid.uuid4())
            cursor.execute(
                "INSERT INTO orders (uuid, nama, created_at) VALUES (?, ?, ?);",
                (order_uuid, nama_raw, row_dict.get("created_at", now)),
            )
            orders_map[nama_key] = cursor.lastrowid

        order_id = orders_map[nama_key]

        # Insert item
        item_uuid = str(uuid.uuid4())
        cursor.execute(
            """
            INSERT INTO ritual_items
                (uuid, order_id, panggilan, nama_mandarin,
                 penyebutan, dari, keluarga, keterangan,
                 tahun_lunar, bulan_lunar, hari_lunar, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
            """,
            (
                item_uuid,
                order_id,
                row_dict.get("panggilan", ""),
                row_dict.get("nama_mandarin", ""),
                row_dict.get("penyebutan", ""),
                row_dict.get("dari", ""),
                row_dict.get("keluarga", ""),
                row_dict.get("keterangan", ""),
                row_dict.get("tahun_lunar", ""),
                row_dict.get("bulan_lunar", ""),
                row_dict.get("hari_lunar", ""),
                row_dict.get("created_at", now),
            ),
        )

    # Rename tabel lama → backup
    cursor.execute("ALTER TABLE ritual_forms RENAME TO ritual_forms_backup;")
    conn.commit()


# ============================================================
# CRUD — Orders
# ============================================================
def get_or_create_order(
    nama: str, db_path: str = _DB_PATH
) -> tuple[int, str]:
    """
    Ambil order yang sudah ada berdasarkan nama (case-insensitive),
    atau buat baru jika belum ada.

    Args:
        nama:    Nama pemesan (Indonesia).
        db_path: Path ke database.

    Returns:
        tuple[int, str]: (order_id, order_uuid)
    """
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()

        # Cari order yang sudah ada
        cursor.execute(
            "SELECT id, uuid FROM orders WHERE LOWER(TRIM(nama)) = LOWER(TRIM(?));",
            (nama,),
        )
        row = cursor.fetchone()
        if row:
            return row["id"], row["uuid"]

        # Buat order baru
        order_uuid = str(uuid.uuid4())
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cursor.execute(
            "INSERT INTO orders (uuid, nama, created_at) VALUES (?, ?, ?);",
            (order_uuid, nama.strip(), now),
        )
        conn.commit()
        return cursor.lastrowid, order_uuid

    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal get/create order: {e}") from e
    finally:
        conn.close()


def get_all_orders(db_path: str = _DB_PATH) -> list[dict]:
    """
    Ambil semua order, beserta jumlah item per order.

    Returns:
        list[dict]: [{id, uuid, nama, item_count, created_at}, ...]
    """
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT o.id, o.uuid, o.nama, o.created_at,
                   COUNT(ri.id) AS item_count
            FROM orders o
            LEFT JOIN ritual_items ri ON ri.order_id = o.id
            GROUP BY o.id
            ORDER BY o.nama COLLATE NOCASE, o.created_at DESC;
        """)
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal mengambil data orders: {e}") from e
    finally:
        conn.close()


def delete_order(order_uuid: str, db_path: str = _DB_PATH) -> bool:
    """
    Hapus order beserta semua item-nya (CASCADE).

    Returns:
        bool: True jika berhasil dihapus.
    """
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM orders WHERE uuid = ?;", (order_uuid,))
        conn.commit()
        return cursor.rowcount > 0
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal menghapus order: {e}") from e
    finally:
        conn.close()


def update_order_name(
    order_uuid: str, new_name: str, db_path: str = _DB_PATH,
) -> dict:
    """
    Rename nama pemesan pada order.

    Jika sudah ada order lain dengan nama yang sama, semua item dari
    order lama akan dipindah ke order yang sudah ada, lalu order lama dihapus.

    Args:
        order_uuid: UUID order yang akan di-rename.
        new_name:   Nama baru.

    Returns:
        dict: {"success": True, "merged": bool}
    """
    new_name = new_name.strip()
    if not new_name:
        raise RuntimeError("Nama pemesan tidak boleh kosong.")
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()

        # Cek apakah sudah ada order dengan nama yang sama
        cursor.execute(
            "SELECT id FROM orders WHERE LOWER(TRIM(nama)) = LOWER(TRIM(?)) "
            "AND uuid != ?;",
            (new_name, order_uuid),
        )
        existing = cursor.fetchone()

        merged = False
        if existing:
            # Pindahkan semua item ke order yang sudah ada
            target_id = existing["id"]
            cursor.execute(
                "SELECT id FROM orders WHERE uuid = ?;", (order_uuid,),
            )
            old_row = cursor.fetchone()
            if old_row:
                cursor.execute(
                    "UPDATE ritual_items SET order_id = ? WHERE order_id = ?;",
                    (target_id, old_row["id"]),
                )
                cursor.execute(
                    "DELETE FROM orders WHERE uuid = ?;", (order_uuid,),
                )
                merged = True
        else:
            cursor.execute(
                "UPDATE orders SET nama = ? WHERE uuid = ?;",
                (new_name, order_uuid),
            )
        conn.commit()
        return {"success": True, "merged": merged}
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal rename order: {e}") from e
    finally:
        conn.close()


def bulk_delete_orders(
    order_uuids: list[str], db_path: str = _DB_PATH,
) -> int:
    """
    Hapus beberapa order sekaligus beserta semua item-nya.

    Args:
        order_uuids: List UUID order yang akan dihapus.

    Returns:
        int: Jumlah order yang berhasil dihapus.
    """
    if not order_uuids:
        return 0
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()
        placeholders = ",".join("?" for _ in order_uuids)
        cursor.execute(
            f"DELETE FROM orders WHERE uuid IN ({placeholders});",
            order_uuids,
        )
        conn.commit()
        return cursor.rowcount
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal bulk delete orders: {e}") from e
    finally:
        conn.close()


# ============================================================
# CRUD — Ritual Items
# ============================================================
def insert_record(
    panggilan: str,
    nama_mandarin: str,
    dari: str,
    nama: str = "",
    penyebutan: str = "",
    keluarga: str = "",
    keterangan: str = "",
    tahun_lunar: str = "",
    bulan_lunar: str = "",
    hari_lunar: str = "",
    db_path: str = _DB_PATH,
) -> str:
    """
    Menyimpan satu record baru.
    Otomatis membuat/mengambil order berdasarkan nama.

    Args:
        panggilan:     Sebutan Mandarin (母親許門, 父親, 祖父).
        nama_mandarin: Nama Mandarin mendiang (梁氏橋玉).
        dari:          Pengirim / hubungan (孝男, 外孫女敬奉).
        nama:          Nama pemesan Indonesia.
        penyebutan:    Romanisasi / Hokkien.
        keluarga:      Relasi keluarga Indonesia.
        keterangan:    Info tambahan.
        tahun_lunar:   Tahun lunar.
        bulan_lunar:   Bulan lunar.
        hari_lunar:    Hari lunar.
        db_path:       Path ke database.

    Returns:
        str: UUID dari item yang baru dibuat.
    """
    # Get atau create order berdasarkan nama
    order_id, _ = get_or_create_order(nama, db_path)

    item_uuid = str(uuid.uuid4())
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO ritual_items
                (uuid, order_id, panggilan, nama_mandarin,
                 penyebutan, dari, keluarga, keterangan,
                 tahun_lunar, bulan_lunar, hari_lunar, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
            """,
            (
                item_uuid, order_id, panggilan, nama_mandarin,
                penyebutan, dari, keluarga, keterangan,
                tahun_lunar, bulan_lunar, hari_lunar, now,
            ),
        )
        conn.commit()
        return item_uuid
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal menyimpan data: {e}") from e
    finally:
        conn.close()


def update_record(
    record_uuid: str,
    panggilan: str = "",
    nama_mandarin: str = "",
    dari: str = "",
    nama: str = "",
    penyebutan: str = "",
    keluarga: str = "",
    keterangan: str = "",
    db_path: str = _DB_PATH,
) -> bool:
    """
    Update satu record ritual_item berdasarkan UUID.
    Jika nama pemesan berubah, item dipindah ke order baru/lama
    dan order lama dihapus jika kosong.

    Args:
        record_uuid:   UUID item yang akan diupdate.
        panggilan:     Sebutan Mandarin.
        nama_mandarin: Nama Mandarin mendiang.
        dari:          Pengirim / hubungan.
        nama:          Nama pemesan Indonesia (boleh berubah).
        penyebutan:    Romanisasi / Hokkien.
        keluarga:      Relasi keluarga Indonesia.
        keterangan:    Info tambahan.
        db_path:       Path ke database.

    Returns:
        bool: True jika berhasil diupdate.
    """
    # Get atau create order berdasarkan nama (koneksi terpisah)
    new_order_id, _ = get_or_create_order(nama, db_path)

    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()

        # Ambil order_id lama
        cursor.execute(
            "SELECT order_id FROM ritual_items WHERE uuid = ?;",
            (record_uuid,),
        )
        row = cursor.fetchone()
        if not row:
            return False

        old_order_id = row["order_id"]

        # Update item
        cursor.execute(
            """
            UPDATE ritual_items
            SET order_id = ?, panggilan = ?, nama_mandarin = ?,
                penyebutan = ?, dari = ?, keluarga = ?,
                keterangan = ?
            WHERE uuid = ?;
            """,
            (
                new_order_id, panggilan, nama_mandarin,
                penyebutan, dari, keluarga, keterangan,
                record_uuid,
            ),
        )

        # Hapus order lama jika kosong setelah pindah
        if old_order_id != new_order_id:
            cursor.execute(
                "SELECT COUNT(*) AS cnt FROM ritual_items WHERE order_id = ?;",
                (old_order_id,),
            )
            if cursor.fetchone()["cnt"] == 0:
                cursor.execute(
                    "DELETE FROM orders WHERE id = ?;", (old_order_id,),
                )

        conn.commit()
        return True
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal mengupdate data: {e}") from e
    finally:
        conn.close()


def get_all_records(db_path: str = _DB_PATH) -> list[dict]:
    """
    Mengambil seluruh item + info order, terkelompok per nama pemesan.

    Returns:
        list[dict]: Setiap dict berisi field item + 'nama' dari order.
    """
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ri.uuid, o.nama, ri.panggilan, ri.nama_mandarin,
                   ri.penyebutan, ri.dari, ri.keluarga, ri.keterangan,
                   ri.tahun_lunar, ri.bulan_lunar, ri.hari_lunar,
                   ri.created_at, o.uuid AS order_uuid
            FROM ritual_items ri
            JOIN orders o ON o.id = ri.order_id
            ORDER BY o.nama COLLATE NOCASE, ri.created_at;
        """)
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal mengambil data: {e}") from e
    finally:
        conn.close()


def get_items_by_order(order_uuid: str, db_path: str = _DB_PATH) -> list[dict]:
    """
    Ambil semua item milik satu order.
    """
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ri.uuid, o.nama, ri.panggilan, ri.nama_mandarin,
                   ri.penyebutan, ri.dari, ri.keluarga, ri.keterangan,
                   ri.tahun_lunar, ri.bulan_lunar, ri.hari_lunar,
                   ri.created_at, o.uuid AS order_uuid
            FROM ritual_items ri
            JOIN orders o ON o.id = ri.order_id
            WHERE o.uuid = ?
            ORDER BY ri.created_at;
        """, (order_uuid,))
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal mengambil data: {e}") from e
    finally:
        conn.close()


def get_record_by_uuid(record_uuid: str, db_path: str = _DB_PATH) -> dict | None:
    """
    Mengambil satu item berdasarkan UUID.

    Returns:
        dict | None
    """
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT ri.uuid, o.nama, ri.panggilan, ri.nama_mandarin,
                   ri.penyebutan, ri.dari, ri.keluarga, ri.keterangan,
                   ri.tahun_lunar, ri.bulan_lunar, ri.hari_lunar,
                   ri.created_at, o.uuid AS order_uuid
            FROM ritual_items ri
            JOIN orders o ON o.id = ri.order_id
            WHERE ri.uuid = ?;
        """, (record_uuid,))
        row = cursor.fetchone()
        return dict(row) if row else None
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal mengambil data: {e}") from e
    finally:
        conn.close()


def delete_record(record_uuid: str, db_path: str = _DB_PATH) -> bool:
    """
    Menghapus satu item berdasarkan UUID.
    Jika order-nya menjadi kosong (0 item), order juga dihapus.

    Returns:
        bool: True jika berhasil dihapus.
    """
    try:
        conn = _get_connection(db_path)
        cursor = conn.cursor()

        # Cari order_id sebelum hapus
        cursor.execute(
            "SELECT order_id FROM ritual_items WHERE uuid = ?;",
            (record_uuid,),
        )
        row = cursor.fetchone()
        if not row:
            return False

        order_id = row["order_id"]

        # Hapus item
        cursor.execute("DELETE FROM ritual_items WHERE uuid = ?;", (record_uuid,))

        # Cek apakah order masih punya item lain
        cursor.execute(
            "SELECT COUNT(*) AS cnt FROM ritual_items WHERE order_id = ?;",
            (order_id,),
        )
        if cursor.fetchone()["cnt"] == 0:
            cursor.execute("DELETE FROM orders WHERE id = ?;", (order_id,))

        conn.commit()
        return True
    except sqlite3.Error as e:
        raise RuntimeError(f"Gagal menghapus data: {e}") from e
    finally:
        conn.close()


# ============================================================
# Export Database → Excel (Backup)
# ============================================================
def export_to_excel(
    output_path: str,
    db_path: str = _DB_PATH,
) -> tuple[int, str]:
    """Export seluruh data dari database ke file Excel sebagai backup.

    File yang dihasilkan kompatibel dengan ``import_from_excel()``
    (format legacy — data di Sheet 2).  Sheet 1 berisi ringkasan/info.

    Args:
        output_path: Path tujuan file .xlsx.
        db_path:     Path ke database.

    Returns:
        tuple[int, str]: (jumlah_record, path_file)
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
    except ImportError as e:
        raise RuntimeError(
            "Library openpyxl belum terinstal. "
            "Jalankan: pip install openpyxl"
        ) from e

    records = get_all_records(db_path)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Info Backup ────────────────────────────────
    ws_info = wb.active
    ws_info.title = "Info Backup"

    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B5E20")
    info_font = Font(size=12)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws_info.column_dimensions["A"].width = 30
    ws_info.column_dimensions["B"].width = 50

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    info_rows = [
        ("BACKUP DATABASE RITUAL", ""),
        ("", ""),
        ("Tanggal Backup", now_str),
        ("Jumlah Record", str(len(records))),
        ("Sumber Database", os.path.basename(db_path)),
        ("", ""),
        ("CATATAN", ""),
        (
            "File ini dapat di-import kembali menggunakan",
            'tombol "Import Excel" di aplikasi.',
        ),
        (
            "Jangan mengubah struktur kolom pada sheet",
            '"Data Backup" agar import berjalan lancar.',
        ),
    ]
    for r_idx, (col_a, col_b) in enumerate(info_rows, start=1):
        cell_a = ws_info.cell(row=r_idx, column=1, value=col_a)
        cell_b = ws_info.cell(row=r_idx, column=2, value=col_b)
        cell_a.font = info_font
        cell_b.font = info_font
    # Judul
    ws_info["A1"].font = Font(bold=True, size=16, color="1B5E20")

    # ── Sheet 2: Data Backup ────────────────────────────────
    ws_data = wb.create_sheet("Data Backup")

    headers = [
        "NO", "NAMA PEMESAN", "PANGGILAN",
        "ATAS NAMA MANDARIN", "PENYEBUTAN",
        "DARI", "KELUARGA", "KETERANGAN",
    ]
    col_widths = [6, 25, 20, 25, 20, 25, 20, 25]

    # Header row
    for c_idx, (hdr, width) in enumerate(
        zip(headers, col_widths), start=1,
    ):
        cell = ws_data.cell(row=1, column=c_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        col_letter = openpyxl.utils.get_column_letter(c_idx)
        ws_data.column_dimensions[col_letter].width = width

    # Data rows
    data_font = Font(size=11)
    for r_idx, rec in enumerate(records, start=2):
        row_values = [
            r_idx - 1,
            rec.get("nama", ""),
            rec.get("panggilan", ""),
            rec.get("nama_mandarin", ""),
            rec.get("penyebutan", ""),
            rec.get("dari", ""),
            rec.get("keluarga", ""),
            rec.get("keterangan", ""),
        ]
        for c_idx, val in enumerate(row_values, start=1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    # Freeze header
    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = (
        f"A1:H{max(len(records) + 1, 2)}"
    )

    wb.save(output_path)
    wb.close()

    return (len(records), output_path)


# ============================================================
# Import Excel → Database
# ============================================================


def _auto_keluarga_from_mandarin(panggilan_mandarin: str) -> str:
    """Reverse-lookup aksara Mandarin panggilan → keluarga otomatis.

    Menggunakan data PANGGILAN_OPTIONS dari main module.
    Mendukung item bernomor (template '亡胞__兄' dll).

    Returns:
        String keluarga, atau "" jika tidak ditemukan.
    """
    try:
        from main import (
            PANGGILAN_OPTIONS, _NUMBERED_CATEGORIES,
            _MANDARIN_NUMBERS, _MANDARIN_NUMBERS_ADIK,
        )
    except ImportError:
        return ""

    if not panggilan_mandarin:
        return ""

    # Coba cocokkan item non-bernomor langsung
    for indo, mandarin, keluarga in PANGGILAN_OPTIONS:
        if indo in _NUMBERED_CATEGORIES:
            continue
        if mandarin == panggilan_mandarin:
            return keluarga

    # Coba cocokkan template bernomor
    for indo, template, keluarga in PANGGILAN_OPTIONS:
        if indo not in _NUMBERED_CATEGORIES:
            continue
        parts = template.split("__")
        if len(parts) != 2:
            continue
        prefix, suffix = parts
        if not panggilan_mandarin.startswith(prefix):
            continue
        if suffix and not panggilan_mandarin.endswith(suffix):
            continue
        end_idx = (
            len(panggilan_mandarin) - len(suffix)
            if suffix else len(panggilan_mandarin)
        )
        cn_num = panggilan_mandarin[len(prefix):end_idx]
        if not cn_num:
            continue
        is_adik = "Adik" in indo
        lookup = _MANDARIN_NUMBERS_ADIK if is_adik else _MANDARIN_NUMBERS
        for num, char in lookup.items():
            if char == cn_num:
                return f"{keluarga} {num}"

    return ""


def _auto_penyebutan_from_mandarin(nama_mandarin: str) -> str:
    """Konversi aksara Mandarin → Pinyin (penyebutan) otomatis.

    Returns:
        String Pinyin dengan huruf kapital per suku kata, atau "" jika gagal.
    """
    if not nama_mandarin:
        return ""
    try:
        from pypinyin import pinyin, Style as PinyinStyle
        syllables = pinyin(nama_mandarin, style=PinyinStyle.TONE)
        return " ".join(s[0].capitalize() for s in syllables if s[0])
    except (ImportError, Exception):  # noqa: BLE001
        return ""


def import_from_excel(
    excel_path: str,
    db_path: str = _DB_PATH,
) -> int:
    """
    Import data dari file Excel ke database.

    Format didukung:
      1. **Template baru** — Sheet "Data Import" (index 0).
         Kolom: A=NO, B=NAMA, C=PANGGILAN (Indonesia),
                D=ATAS NAMA MANDARIN, E=PENYEBUTAN,
                F=DARI (Indonesia), G=KETERANGAN.
         Panggilan & Dari otomatis dikonversi ke aksara Mandarin.

      2. **Backup** — Sheet "Data Backup".
         Kolom: A=NO, B=NAMA, C=PANGGILAN (Mandarin),
                D=ATAS NAMA MANDARIN, E=PENYEBUTAN,
                F=DARI (Mandarin), G=KELUARGA, H=KETERANGAN.

      3. **Format lama** — Sheet 2 (index 1).
         Kolom sama seperti format Backup.

    Args:
        excel_path:  Path ke file .xlsx
        db_path:     Path ke database.

    Returns:
        int: Jumlah record yang berhasil diimport.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "Library openpyxl belum terinstal. "
            "Jalankan: pip install openpyxl"
        ) from e

    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)

        # Deteksi format berdasarkan nama sheet
        is_new_format = "Data Import" in wb.sheetnames
        is_backup = "Data Backup" in wb.sheetnames

        if is_new_format:
            ws = wb["Data Import"]
            # Lazy import konverter
            try:
                from modules.excel_template import (
                    convert_panggilan, convert_dari,
                )
            except ImportError:
                convert_panggilan = None  # type: ignore[assignment]
                convert_dari = None  # type: ignore[assignment]
        elif is_backup:
            ws = wb["Data Backup"]
            convert_panggilan = None  # type: ignore[assignment]
            convert_dari = None  # type: ignore[assignment]
        else:
            ws = wb.worksheets[1]  # Sheet 2 (legacy)
            convert_panggilan = None  # type: ignore[assignment]
            convert_dari = None  # type: ignore[assignment]

        # Kumpulkan semua data dulu, baru insert ke DB
        rows_data: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            nama = str(row[1] or "").strip() if len(row) > 1 else ""
            panggilan_raw = str(row[2] or "").strip() if len(row) > 2 else ""
            nama_mandarin = str(row[3] or "").strip() if len(row) > 3 else ""
            penyebutan = str(row[4] or "").strip() if len(row) > 4 else ""
            dari_raw = str(row[5] or "").strip() if len(row) > 5 else ""

            # Skip baris kosong (tanpa panggilan DAN nama mandarin)
            if not panggilan_raw and not nama_mandarin:
                continue

            # Konversi Indonesia → Mandarin (template baru)
            keluarga = ""
            if is_new_format and convert_panggilan:
                result = convert_panggilan(panggilan_raw)
                if result:
                    panggilan, keluarga = result
                else:
                    panggilan = panggilan_raw
            else:
                panggilan = panggilan_raw

            if is_new_format and convert_dari:
                dari = convert_dari(dari_raw) or dari_raw
            else:
                dari = dari_raw

            # Kolom keluarga & keterangan (posisi beda per format)
            if not is_new_format:
                keluarga = (
                    str(row[6] or "").strip() if len(row) > 6 else ""
                )
                keterangan = (
                    str(row[7] or "").strip() if len(row) > 7 else ""
                )
            else:
                keterangan = (
                    str(row[6] or "").strip() if len(row) > 6 else ""
                )

            # --- Auto-fill keluarga dari panggilan Mandarin ---
            if not keluarga and panggilan:
                keluarga = _auto_keluarga_from_mandarin(panggilan)

            # --- Auto-fill penyebutan dari nama Mandarin (Pinyin) ---
            if not penyebutan and nama_mandarin:
                penyebutan = _auto_penyebutan_from_mandarin(nama_mandarin)

            rows_data.append({
                "panggilan": panggilan,
                "nama_mandarin": nama_mandarin,
                "dari": dari,
                "nama": nama,
                "penyebutan": penyebutan,
                "keluarga": keluarga,
                "keterangan": keterangan,
            })

        # Tutup workbook SEGERA setelah selesai baca
        wb.close()
        wb = None

        # Import baru: insert ke DB setelah file Excel ditutup
        import gc
        gc.collect()

        count = 0
        for rd in rows_data:
            insert_record(
                panggilan=rd["panggilan"],
                nama_mandarin=rd["nama_mandarin"],
                dari=rd["dari"],
                nama=rd["nama"],
                penyebutan=rd["penyebutan"],
                keluarga=rd["keluarga"],
                keterangan=rd["keterangan"],
                db_path=db_path,
            )
            count += 1

        return count

    except (OSError, KeyError) as e:
        raise RuntimeError(f"Gagal membaca file Excel: {e}") from e
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"Gagal import Excel: {e}") from e
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


# ============================================================
# Jika dijalankan langsung, inisialisasi database sebagai test.
# ============================================================
if __name__ == "__main__":
    init_db()
    print(f"Database berhasil diinisialisasi di: {_DB_PATH}")

    # Test: tampilkan orders
    orders = get_all_orders()
    print(f"\n=== Orders ({len(orders)}) ===")
    for o in orders:
        print(f"  [{o['uuid'][:8]}] {o['nama']} — {o['item_count']} item(s)")

    # Test: tampilkan semua items grouped
    records = get_all_records()
    current_nama = None
    for r in records:
        if r["nama"] != current_nama:
            current_nama = r["nama"]
            print(f"\n--- {current_nama or '(tanpa nama)'} ---")
        print(f"  {r['panggilan']} | {r['nama_mandarin']} | {r['dari']}")
